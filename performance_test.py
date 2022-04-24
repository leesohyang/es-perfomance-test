import os
import json
from elasticsearch import Elasticsearch, helpers
import uuid
import numpy as np
from openpyxl import load_workbook
import re
import time
import requests
from multiprocessing import Process, Queue
import multiprocessing
import sys
from nested_lookup import nested_lookup
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as tic
from jsonpath_ng import jsonpath, parse
from datetime import datetime
import logging
import logging.handlers
import os
from pprint import pprint
from openpyxl import load_workbook

# Create logger instance & setup to log level
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Create to formater
formatter = logging.Formatter('[%(levelname)s|%(filename)s:%(lineno)s] %(asctime)s > %(message)s')

# Create to fileHandler and StreamHandler
fileHandler = logging.FileHandler('result.log')
streamHandler = logging.StreamHandler()

# Setup to formater on Handler
fileHandler.setFormatter(formatter)
streamHandler.setFormatter(formatter)

# Add to Handler on logging
logger.addHandler(fileHandler)
logger.addHandler(streamHandler)

# logging
# logging.debug("debug")
logging.info("info")
# logging.warning("warning")
# logging.error("error")
# logging.critical("critical")

# result logging
# sys.stdout = open('result.log', 'w')


result_memory = list()
columns = ['timestamp', 'cpu', 'jvm', 'index_total', 'index_time_in_millis', 'io_write', 'io_read', 'query_total',
           'query_time_in_millis']


def append_df_to_excel(filename, df, sheet_name='Sheet1', start_row=None, truncate_sheet=False, **kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      start_row : upper left cell row to dump data frame.
                 Per default (start_row=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in kwargs:
        kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if start_row is None and sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if start_row is None:
        start_row = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=start_row, index=False, header=False, **kwargs)  # header=True

    # save the workbook
    writer.save()


def monitor(event):
    elastic = Elasticsearch(hosts="", timeout=30, max_retris=10, retry_on_timeout=True)
    index = 'edgengram'
    df = pd.DataFrame(columns=columns)
    start_time_m = time.time()
    while True:
        row = list()
        stats = elastic.nodes.stats()
        timestamp = [match.value for match in parse('$..EmKP3BpTRZCoau5TiHQPBA.timestamp').find(stats)]
        cpu = [match.value for match in parse('$..process.cpu.percent').find(stats)]
        jvm = [match.value for match in parse('$..heap_used_in_bytes').find(stats)]
        index_total = [match.value for match in parse('$..index_total').find(stats)]
        index_time_in_millis = [match.value for match in parse('$..index_time_in_millis').find(stats)]
        io_write = [match.value for match in parse('$..total.write_operations').find(stats)]
        io_read = [match.value for match in parse('$..total.read_operations').find(stats)]
        query_total = [match.value for match in parse('$..query_total').find(stats)]
        query_time_in_millis = [match.value for match in parse('$..query_time_in_millis').find(stats)]

        df = pd.DataFrame([[datetime.fromtimestamp(timestamp[0] / 1e3).strftime('%Y-%m-%d %H:%M:%S'), cpu[0], jvm[0],
                            index_total[0], index_time_in_millis[0], io_write[0], io_read[0], query_total[0],
                            query_time_in_millis[0]]], columns=columns)

        time.sleep(30)  # 30초마다 check

        append_df_to_excel("second_test_t.xlsx", df)


def producer(queue):
    elastic = Elasticsearch(hosts="", timeout=30, max_retris=10, retry_on_timeout=True,
                            maxsize=25)
    index = 'test'
       res = elastic.search(index=index,
           body={
               "query":{
                   "exists":{ "field":"rawData"}
               }
           },
           scroll='10m', size=10000)
       sid = res['_scroll_id']
       total_cnt = res['hits']['hits']
       scroll_size = len(res['hits']['hits'])

       scroll_count = 0
       while scroll_size > 0:
           if queue.empty():
               scroll_count += 1
               for data in res['hits']['hits']:
                   queue.put(data)

               res = elastic.scroll(scroll_id=sid, scroll='2m')
               sid = res['_scroll_id']
               scroll_size = len(res['hits']['hits'])

def searcher(queue):
    elastic = Elasticsearch(hosts="", timeout=30, max_retris=10, retry_on_timeout=True,
                            maxsize=25)
    index = 'test'
    r = requests.post("http:///edgengram/_cache/clear")

    while True:
       if queue.empty():
           continue
       data = queue.get()
       logger.info(f"get search term with process ID: {os.getpid()}")
       try:
           term = data['_source']['rawData']
           term_vector = list()
           term_vector.append(term)
           for term in term_vector:
               term = " ".join(re.findall("[a-zA-Z0-9가-핳]+", term))[-10:]
               logger.info(f"search term is {term}")
               result = elastic.search(
                   index="test",
                   body={
                       "query": {
                           "bool":
                               {"query_string": {
                                   "query": term, "default_field": "rawData"
                               }}
                       }
                   }
               )
       except KeyError:
           continue

    # if time.time() - start_times > 3600:
        #     event.set()
        # ngram contain search query
        # print(result)


if __name__ == '__main__':
    eventm = multiprocessing.Event()
    events = multiprocessing.Event()
    start_time = time.time()
    p1 = Process(target=monitor, args=(eventm,))
    MAX_CONNECTS = 25

    for i in range(1, 6):
        queue = Queue()
        p2 = Process(target=producer, args=(queue,))
        p2.start()
        jobs = []
        for j in range(5 * i):
            logger.info("start searcher")
            p3 = Process(target=searcher, args=(queue,))
            p3.start()
            jobs.append(p3)
        queue.close()
        queue.join_thread()
        for p in range(len(jobs)):
            jobs[p].join()
        time.sleep(3600)
        for job in jobs:
            job.terminate()

    p1.terminate()
    p2.terminate()
    sys.exit()

    # while True:
    #     if flag == False and time.time() - start_time >= 3600:
    #         flag = True
    #         p2 = Process(target=searcher, args=(events,))
    #         print("start searcher")
    #         p2.start()  # after 20min, searcher start
    #     if events.is_set():
    #         # print("Exiting searcher")
    #         p2.terminate()
    #
    #     if eventm.is_set():
    #         print("Exiting monitoring")
    #         p1.terminate()
    #         sys.exit()